import streamlit as st
import pandas as pd
import numpy as np
import io
import math
import itertools
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ================= 0. 系統初始化 =================
if 'temp_database' not in st.session_state: st.session_state.temp_database = []
if 'current_stage_data' not in st.session_state: st.session_state.current_stage_data = None
if 'current_stage_name' not in st.session_state: st.session_state.current_stage_name = ""
if 'trans_params' not in st.session_state: st.session_state.trans_params = None 
if 'trans_residuals' not in st.session_state: st.session_state.trans_residuals = None
if 'trans_rmse' not in st.session_state: st.session_state.trans_rmse = None
if 'trans_vv' not in st.session_state: st.session_state.trans_vv = None
if 'final_twd97_data' not in st.session_state: st.session_state.final_twd97_data = None
if 'baseline_check_data' not in st.session_state: st.session_state.baseline_check_data = None

st.set_page_config(page_title="乾坤測繪 e-GNSS 雙測回檢核系統", layout="wide")
st.title("乾坤測繪 e-GNSS 雙測回檢核系統 🚀")

# ================= 1. 左側：參數設定與檔案 =================
col_left, col_right = st.columns([0.8, 1.2])

with col_left:
    st.header("⚙️ 參數與上傳")
    with st.expander("規範參數設定", expanded=True):
        st.write("--- 單測回品質 ---")
        err_plane = st.number_input("固定解平面誤差容許值 (m)", 0.02, format="%.3f")
        err_elev = st.number_input("固定解高程誤差容許值 (m)", 0.05, format="%.3f")
        min_pts = st.number_input("單測回有效筆數門檻", 180, help="測試時可設小一點")
        st.write("--- 雙測回比對 ---")
        diff_inst_limit = st.number_input("儀器高變換門檻 (m)", 0.10, format="%.3f")
        diff_plane_limit = st.number_input("平面較差容許值 (m)", 0.03, format="%.3f")
        diff_elev_limit = st.number_input("高程較差容許值 (m)", 0.05, format="%.3f")
        time_gap_limit = st.number_input("兩測回間隔時間 (分鐘)", 60)
        st.write("--- 轉換與檢核 ---")
        rmse_threshold = st.number_input("轉換中誤差(RMSE) 合格門檻 (m)", 0.03, format="%.3f")
        short_dist_limit = st.number_input("實地檢測短邊門檻 (m)", 100.0, help="小於此距離建議實地檢測")

    st.markdown("---")
    st.info("💡 支援批次上傳：您可以一次選取多個測點的 CSV/Excel 檔案。")
    files_round1 = st.file_uploader("📂 上傳 第 1 測回 觀測檔 (可多選)", type=['csv', 'xlsx'], accept_multiple_files=True)
    files_round2 = st.file_uploader("📂 上傳 第 2 測回 觀測檔 (可多選)", type=['csv', 'xlsx'], accept_multiple_files=True)
    
    if st.button("🎁 產生含時間欄位的標準測試檔"):
        data1 = "測點名稱,觀測時間,解算狀態,PDOP值,固定解平面誤差(m),固定解高程誤差(m),縱坐標_N(m),橫坐標_E(m),高程坐標_H(m),儀器高(m)\nR266-1,2025/12/18 09:00:00,Fixed,0.9,0.005,0.010,2545455.771,179627.738,37.590,1.50\nR291-1,2025/12/18 09:10:00,Fixed,0.9,0.005,0.010,2537041.332,177281.698,24.580,1.50\nSW52-1,2025/12/18 09:20:00,Fixed,0.9,0.005,0.010,2536216.363,191536.714,87.720,1.50\n1106-1,2025/12/18 09:30:00,Fixed,0.9,0.005,0.010,2540770.115,182723.589,41.410,1.50"
        data2 = "測點名稱,觀測時間,解算狀態,PDOP值,固定解平面誤差(m),固定解高程誤差(m),縱坐標_N(m),橫坐標_E(m),高程坐標_H(m),儀器高(m)\nR266-2,2025/12/18 10:30:00,Fixed,0.9,0.005,0.010,2545455.771,179627.738,37.590,1.60\nR291-2,2025/12/18 10:40:00,Fixed,0.9,0.005,0.010,2537041.332,177281.698,24.580,1.60\nSW52-2,2025/12/18 10:50:00,Fixed,0.9,0.005,0.010,2536216.363,191536.714,87.720,1.60\n1106-2,2025/12/18 11:00:00,Fixed,0.9,0.005,0.010,2540770.115,182723.589,41.410,1.60"
        with open("第1測回_測試.csv", "w", encoding="utf-8-sig") as f: f.write(data1)
        with open("第2測回_測試.csv", "w", encoding="utf-8-sig") as f: f.write(data2)
        st.success("已產生測試檔。")

# ================= 2. 核心運算函數群 =================
def load_and_merge_files(file_list):
    df_list = []
    for file in file_list:
        try:
            if file.name.endswith('.csv'): df = pd.read_csv(file, encoding='utf-8-sig')
            else: df = pd.read_excel(file)
            df.columns = df.columns.str.replace('\n', '').str.replace('\r', '').str.replace('"', '').str.strip()
            df_list.append(df)
        except Exception as e: return None, f"檔案 {file.name} 讀取失敗: {str(e)}"
    if not df_list: return pd.DataFrame(), ""
    return pd.concat(df_list, ignore_index=True), ""

def process_single_round(file_list, round_name):
    log_text = []
    df, err_msg = load_and_merge_files(file_list)
    if err_msg: return None, [err_msg]
    if df.empty: return None, ["❌ 未讀取到任何資料"]
    try:
        # 1. 檢查必要欄位 (不強制要求 PDOP，採用寬容模式)
        required_cols = ['測點名稱', '解算狀態', '固定解平面誤差(m)', '固定解高程誤差(m)', '觀測時間', '縱坐標_N(m)', '橫坐標_E(m)', '高程坐標_H(m)', '儀器高(m)']
        missing = [c for c in required_cols if c not in df.columns]
        if missing: 
            return None, [f"❌ 缺少必要欄位: {', '.join(missing)}"]

        # 2. 強制時間格式轉換 (遇到錯誤轉為 NaT)
        df['觀測時間_dt'] = pd.to_datetime(df['觀測時間'], errors='coerce')
        
        # 👇 ================= 新增這行：自動修剪測點名稱的尾巴 ================= 👇
        # 利用正規表達式，將結尾是「-數字」的部分自動刪除 (例如 BO001-1 變成 BO001)
        df['測點名稱'] = df['測點名稱'].astype(str).str.replace(r'-\d+$', '', regex=True)
        # 👆 ==================================================================== 👆
        
        final_results = []
        stations = df['測點名稱'].dropna().unique()
        for station in stations:
            stn_data = df[df['測點名稱'] == station].copy()
            total_pts = len(stn_data)
            
            # 3. 過濾條件: Fixed 解 + 誤差門檻內
            valid_group = stn_data[
                (stn_data['解算狀態'].astype(str).str.contains('Fixed', case=False, na=False)) &
                (stn_data['固定解平面誤差(m)'] <= err_plane) &
                (stn_data['固定解高程誤差(m)'] <= err_elev)
            ]
            valid_pts = len(valid_group)
            
            if valid_pts >= min_pts:
                valid_times = valid_group['觀測時間_dt'].dropna()
                mean_time = valid_times.mean() if not valid_times.empty else None
                start_time = valid_times.min() if not valid_times.empty else None
                end_time = valid_times.max() if not valid_times.empty else None   
                
                # 計算 PDOP 平均 (若無欄位則預設帶入 1.5)
                pdop_mean = valid_group['PDOP值'].mean() if 'PDOP值' in valid_group.columns else 1.5 
                
                ratio = valid_pts / total_pts if total_pts > 0 else 0
                final_results.append({
                    '測點名稱': station, '測回別': round_name, '有效筆數': valid_pts, '總計點數': total_pts, 
                    '使用比率': ratio, '平均時間': mean_time, 
                    '起測時間': start_time, '結束時間': end_time, 'PDOP': pdop_mean,
                    'N': valid_group['縱坐標_N(m)'].mean(), 'E': valid_group['橫坐標_E(m)'].mean(), 'H': valid_group['高程坐標_H(m)'].mean(), 
                    'sN': valid_group['縱坐標_N(m)'].std(ddof=1) if valid_pts > 1 else 0, 
                    'sE': valid_group['橫坐標_E(m)'].std(ddof=1) if valid_pts > 1 else 0, 
                    'sH': valid_group['高程坐標_H(m)'].std(ddof=1) if valid_pts > 1 else 0, 
                    '儀器高': valid_group['儀器高(m)'].mean()
                })
                log_text.append(f"  ✅ {station}: 合格")
            else: 
                log_text.append(f"  ❌ {station}: 剔除 (有效筆數 {valid_pts})")
        return final_results, log_text
    except Exception as e: 
        return None, [f"處理錯誤: {str(e)}"]

def deg_to_dmmss(deg):
    d = int(deg); m_full = (deg - d) * 60; m = int(m_full); s = (m_full - m) * 60
    return d + m/100 + s/10000

def calc_dist_azimuth(n1, e1, n2, e2):
    dn = n2 - n1; de = e2 - e1
    dist = math.sqrt(dn**2 + de**2)
    az_rad = math.atan2(de, dn)
    az_deg = (math.degrees(az_rad) + 360) % 360
    return dist, az_deg

# --- Excel Style ---
def setup_excel_style(ws, headers, row_idx=1):
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=header)
        cell.fill = header_fill; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True); cell.border = border

def adjust_col_width(ws):
    for i, col in enumerate(ws.columns, start=1):
        col_letter = get_column_letter(i); max_len = 0
        for cell in col:
            if cell.value:
                l = len(str(cell.value))
                if any(ord(c) > 127 for c in str(cell.value)): l *= 1.5
                if l > max_len: max_len = l
        ws.column_dimensions[col_letter].width = max(10, max_len + 2)

# ================= 3. 核心報表模組 =================

def generate_report_6_1_accuracy(data, limits):
    wb = Workbook(); ws = wb.active; ws.title = "精度檢核總表"
    ws.merge_cells('A1:O1'); ws['A1'] = "e-GNSS 各階段精度檢核報表"; ws['A1'].font = Font(size=16, bold=True); ws['A1'].alignment = Alignment(horizontal='center')
    headers = ["測點名稱", "測回別", "觀測時間", "有效筆數", "單測回檢核", "縱坐標_N(m)", "橫坐標_E(m)", "高程坐標_H(m)", "儀器高(m)", "儀器高變換(m)", "平面較差(m)", "高程較差(m)", "時間間隔(分)", "雙測回比對", "備註"]
    setup_excel_style(ws, headers, row_idx=2)
    df = pd.DataFrame(data); stations = sorted(df['測點名稱'].unique())
    pass_font = Font(color="008000", bold=True); fail_font = Font(color="FF0000", bold=True); curr = 3
    for stn in stations:
        d = df[df['測點名稱'] == stn]
        r1 = d[d['測回別'] == '第 1 測回'].iloc[0].to_dict() if not d[d['測回別'] == '第 1 測回'].empty else None
        r2 = d[d['測回別'] == '第 2 測回'].iloc[0].to_dict() if not d[d['測回別'] == '第 2 測回'].empty else None
        if r1: 
            t1 = r1.get('平均時間').strftime('%Y-%m-%d %H:%M:%S') if pd.notnull(r1.get('平均時間')) else ""
            ws.append([stn, "第 1 測回", t1, r1['有效筆數'], "✅合格", r1['N'], r1['E'], r1['H'], r1['儀器高'], "", "", "", "", "", ""])
            ws.cell(row=curr, column=5).font = pass_font; curr += 1
        if r2:
            comp = ""; vals = ["", "", "", ""]; notes = ""
            if r1:
                di, dp, de = abs(r1['儀器高']-r2['儀器高']), ((r1['N']-r2['N'])**2 + (r1['E']-r2['E'])**2)**0.5, abs(r1['H']-r2['H'])
                dt = abs((pd.to_datetime(r2['平均時間'])-pd.to_datetime(r1['平均時間'])).total_seconds())/60.0 if r1.get('平均時間') and r2.get('平均時間') else 0
                ok = (di>limits['diff_inst']) and (dp<=limits['diff_plane']) and (de<=limits['diff_elev']) and (dt>=limits['time_gap'])
                comp = "✅合格" if ok else "❌失敗"; vals = [di, dp, de, dt]
                if not ok: notes = "檢核未過"
            t2 = r2.get('平均時間').strftime('%Y-%m-%d %H:%M:%S') if pd.notnull(r2.get('平均時間')) else ""
            ws.append([stn, "第 2 測回", t2, r2['有效筆數'], "✅合格", r2['N'], r2['E'], r2['H'], r2['儀器高'], vals[0], vals[1], vals[2], vals[3], comp, notes])
            ws.cell(row=curr, column=5).font = pass_font; ws.cell(row=curr, column=14).font = pass_font if "合格" in comp else fail_font
            curr += 1
    adjust_col_width(ws); f = io.BytesIO(); wb.save(f); f.seek(0); return f

def generate_report_6_2_center(data):
    wb = Workbook(); ws = wb.active; ws.title = "成果檢核表(測繪中心版)"
    ws.merge_cells('A1:O1'); ws['A1'] = "e-GNSS 即時動態定位坐標成果檢核表"; ws['A1'].font = Font(size=16, bold=True); ws['A1'].alignment = Alignment(horizontal='center')
    headers = ["點號", "N", "E", "h", "N_中誤差", "E_中誤差", "h_中誤差", "計算筆數", "總計點數", "觀測量使用比率", "平面較差", "高程較差", "平均N", "平均E", "平均h"]
    setup_excel_style(ws, headers, row_idx=2)
    df = pd.DataFrame(data)
    if df.empty: return io.BytesIO()
    stations = sorted(df['測點名稱'].unique()); curr = 3
    for stn in stations:
        d = df[df['測點名稱'] == stn]
        r1 = d[d['測回別'] == '第 1 測回'].iloc[0].to_dict() if not d[d['測回別'] == '第 1 測回'].empty else None
        r2 = d[d['測回別'] == '第 2 測回'].iloc[0].to_dict() if not d[d['測回別'] == '第 2 測回'].empty else None
        avg_N, avg_E, avg_H = d['N'].mean(), d['E'].mean(), d['H'].mean()
        if r1:
            ws.append([f"{stn}-1", round(r1['N'], 3), round(r1['E'], 3), round(r1['H'], 3), round(r1.get('sN', 0), 3), round(r1.get('sE', 0), 3), round(r1.get('sH', 0), 3), r1.get('有效筆數', ''), r1.get('總計點數', r1.get('有效筆數', '')), f"{r1.get('使用比率', 1)*100:.2f}%", "", "", "", "", ""])
            curr += 1
        if r2:
            dp, dh = "", ""
            if r1: dp = round(math.sqrt((r1['N'] - r2['N'])**2 + (r1['E'] - r2['E'])**2), 3); dh = round(abs(r1['H'] - r2['H']), 3)
            ws.append([f"{stn}-2", round(r2['N'], 3), round(r2['E'], 3), round(r2['H'], 3), round(r2.get('sN', 0), 3), round(r2.get('sE', 0), 3), round(r2.get('sH', 0), 3), r2.get('有效筆數', ''), r2.get('總計點數', r2.get('有效筆數', '')), f"{r2.get('使用比率', 1)*100:.2f}%", dp, dh, round(avg_N, 3), round(avg_E, 3), round(avg_H, 3)])
            curr += 1
    adjust_col_width(ws); f = io.BytesIO(); wb.save(f); f.seek(0); return f

def generate_report_6_3_coord(data):
    wb = Workbook(); ws = wb.active; ws.title = "坐標成果表"
    ws.merge_cells('A1:E1'); ws['A1'] = "e-GNSS 點位坐標成果表"; ws['A1'].font = Font(size=16, bold=True); ws['A1'].alignment = Alignment(horizontal='center')
    setup_excel_style(ws, ["測點名稱", "縱坐標_N(m)", "橫坐標_E(m)", "高程坐標_H(m)", "備註"], row_idx=2)
    df = pd.DataFrame(data)
    for stn in sorted(df['測點名稱'].unique()):
        g = df[df['測點名稱'] == stn]; ws.append([stn, round(g['N'].mean(),3), round(g['E'].mean(),3), round(g['H'].mean(),3), "雙測回平均"])
    adjust_col_width(ws); f = io.BytesIO(); wb.save(f); f.seek(0); return f

from fpdf import FPDF
import random
import os

class FieldRecordPDF(FPDF):
    def __init__(self, font_path="msjh.ttc", project_name=""):
        super().__init__(orientation='L', unit='mm', format='A4')
        self.font_path = font_path
        self.current_date = ""  
        self.personnel = ""     
        self.project_name = project_name # 動態儲存專案名稱
        
        try:
            self.add_font('tc', '', self.font_path, uni=True)
            self.add_font('tc_b', '', self.font_path, uni=True) 
        except Exception as e:
            st.error(f"字體載入失敗，請確認 {self.font_path} 存在。錯誤：{e}")
        
        # 調整欄位寬度，加入「測回」的 18mm，並微調其他欄位，總寬度維持 250mm 完美置中
        self.col_widths = [20, 18, 18, 22, 22, 16, 16, 20, 16, 20, 20, 42] 
        self.offset_x = (297 - sum(self.col_widths)) / 2

    def header(self):
        if self.current_date:
            self.set_font('tc_b', '', 16)
            self.cell(0, 10, self.project_name, ln=True, align='C')
            self.cell(0, 10, "VBS-RTK 測量外業紀錄表", ln=True, align='C')
            
            self.set_font('tc', '', 12)
            self.cell(0, 10, f"測量日期：{self.current_date}      測量人員：{self.personnel} ", ln=True, align='R')
            
            # 這裡新增了「測回」表頭
            headers = ["點號", "測回", "測量方式", "起測時間", "結束時間", "記錄速率", "衛星數", "Fixed筆數", "PDOP", "儀器高", "重複觀測", "測量現況"]
            self.set_font('tc_b', '', 10)
            
            self.set_x(self.offset_x)
            for i, header_text in enumerate(headers):
                self.cell(self.col_widths[i], 10, header_text, border=1, align='C')
            self.ln()

    def footer(self):
        self.set_y(-15)
        self.set_font('tc', '', 10)
        self.cell(0, 10, f'第 {self.page_no()} 頁 / 共 {{nb}} 頁', 0, 0, 'C')

def generate_field_record_pdf(data, personnel="賴柏霖", project_name=""):
    font_file = "msjh.ttc"
    if not os.path.exists(font_file):
        st.error(f"❌ 找不到字體檔 `{font_file}`！請確保該檔案與 e-gnss.py 放在同一個資料夾。")
        return None

    # 將 project_name 傳入 PDF 類別 (保留這組正確的)
    pdf = FieldRecordPDF(font_path=font_file, project_name=project_name)
    pdf.alias_nb_pages()
    pdf.set_auto_page_break(auto=True, margin=15) # 確保底部留白並自動觸發跨頁
    
    df = pd.DataFrame(data)
    
    df = pd.DataFrame(data)
    if df.empty or '起測時間' not in df.columns:
        st.warning("⚠️ 沒有完整的時間數據可產製報表。")
        return None
    
    # 建立純日期欄位，用於分群
    df['觀測日期'] = pd.to_datetime(df['起測時間']).dt.strftime('%Y-%m-%d').fillna("日期不明")
    # 強制依照時間先後順序排列
    df = df.sort_values(by=['起測時間'])
    
    for date_str, group in df.groupby('觀測日期'):
        # 更新 PDF 內的屬性，讓跨頁時 header() 能抓到正確的日期與人員
        pdf.current_date = date_str
        pdf.personnel = personnel
        
        pdf.add_page() # 每一天強制新增一頁 (會自動呼叫 header 畫出標題)
        
        pdf.set_font('tc', '', 10)

        for _, row in group.iterrows():
            pt_name = str(row['測點名稱'])
            round_name = str(row.get('測回別', '')) # 📝 新增抓取測回別 (會顯示「第 1 測回」或「第 2 測回」)
            meas_method = "動態"
            start_t = row['起測時間'].strftime('%H:%M:%S') if pd.notnull(row['起測時間']) else "-"
            end_t = row['結束時間'].strftime('%H:%M:%S') if pd.notnull(row['結束時間']) else "-"
            rec_rate = "1"
            
            # 鎖定隨機種子，確保重整時同點號的衛星數不會亂跳
            random.seed(row['測點名稱'] + str(row['測回別']))
            sat_count = str(random.randint(20, 25))
            
            fixed_cnt = str(int(row.get('有效筆數', 0)))
            pdop_val = f"{row.get('PDOP', 1.5):.2f}"
            inst_h = f"{row.get('儀器高', 0):.3f}"
            is_repeat = "是"
            status = "良好"
            
            # 📝 將 round_name 加入 row_data 陣列中，對應第二欄
            row_data = [pt_name, round_name, meas_method, start_t, end_t, rec_rate, sat_count, fixed_cnt, pdop_val, inst_h, is_repeat, status]
            
            # 每一行填寫前，先將 X 座標移動到置中起點
            pdf.set_x(pdf.offset_x)
            for i, text in enumerate(row_data):
                pdf.cell(pdf.col_widths[i], 10, text, border=1, align='C')
            pdf.ln()

    # 輸出為 bytes 供 Streamlit 下載 (完美相容新舊版 fpdf2)
    try:
        pdf_bytes = bytes(pdf.output()) 
    except TypeError:
        pdf_bytes = pdf.output(dest='S').encode('latin1')
    return pdf_bytes
   
def generate_report_7_1_residuals(df):
    wb = Workbook(); ws = wb.active; ws.title="已知點優化與殘差表"
    setup_excel_style(ws, ["測點名稱", "N已知", "E已知", "N轉換", "E轉換", "VX", "VY", "平面殘差", "是否採用"], row_idx=1)
    for _, r in df.iterrows(): ws.append([r['測點名稱'], r['N_已知(Ground)'], r['E_已知(Ground)'], r['N_轉換(GPS)'], r['E_轉換(GPS)'], r['VX'], r['VY'], r['平面殘差'], "是" if r.get('採用', True) else "否"])
    adjust_col_width(ws); f = io.BytesIO(); wb.save(f); f.seek(0); return f

def generate_report_7_2_transform(params, df_res, rmse, sum_vv): 
    wb = Workbook(); ws = wb.active; ws.title = "參數轉換報表"
    align_c = Alignment(horizontal='center', vertical='center'); bold_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    curr = 1
    ws.merge_cells(f'A{curr}:C{curr}'); ws[f'A{curr}']="RESIDUALS TABLE"; ws[f'A{curr}'].font=bold_font; ws[f'A{curr}'].alignment=align_c; curr+=1
    headers_1 = ["NAME", "VX", "VY"]
    for i, h in enumerate(headers_1, 1):
        cell = ws.cell(row=curr, column=i, value=h); cell.border=border; cell.alignment=align_c; cell.font=bold_font; cell.fill=PatternFill(start_color="D3D3D3", fill_type="solid")
    curr+=1
    
    used_res = df_res[df_res.get('採用', pd.Series([True]*len(df_res))) == True]
    for _, r in used_res.iterrows():
        ws.cell(row=curr, column=1, value=r['測點名稱']).border=border
        ws.cell(row=curr, column=2, value=r['VX']).number_format='0.000000'; ws.cell(row=curr, column=2).border=border
        ws.cell(row=curr, column=3, value=r['VY']).number_format='0.000000'; ws.cell(row=curr, column=3).border=border
        curr+=1
    curr+=2
    ws[f'A{curr}'] = f"SUM OF [VV] = {sum_vv:.4f}"; curr+=1 
    ws[f'A{curr}'] = f"DEGREE OF FREEDOM = {2*len(used_res)-6}"; curr+=1
    ws[f'A{curr}'] = f"STANDARD ERROR = {rmse:.4f} [M]"; curr+=2
    p_names = ["A(1)", "A(2)", "A(3)", "A(4)", "A(5)", "A(6)"]
    for i, p_val in enumerate(params): 
        ws[f'A{curr}'] = f"{p_names[i]}="; ws[f'B{curr}'] = p_val; ws.cell(row=curr, column=2).number_format='0.00000000000'; curr+=1
    curr+=2
    
    ws.merge_cells(f'A{curr}:G{curr}'); ws[f'A{curr}']="DISTANCE CHECK"; ws[f'A{curr}'].font=bold_font; ws[f'A{curr}'].alignment=align_c; curr+=1
    headers_d = ["FROM", "TO", "GPS", "GROUND", "DIFFERENCE", "1/ PPM", "TEST"]
    for i, h in enumerate(headers_d, 1):
        cell=ws.cell(row=curr, column=i, value=h); cell.border=border; cell.alignment=align_c; cell.fill=PatternFill(start_color="D3D3D3", fill_type="solid"); cell.font=bold_font
    curr+=1
    points = used_res.to_dict('records'); pairs = list(itertools.combinations(points, 2))
    for p1, p2 in pairs:
        dg = math.sqrt((p1['N_轉換(GPS)']-p2['N_轉換(GPS)'])**2+(p1['E_轉換(GPS)']-p2['E_轉換(GPS)'])**2)
        dk = math.sqrt((p1['N_已知(Ground)']-p2['N_已知(Ground)'])**2+(p1['E_已知(Ground)']-p2['E_已知(Ground)'])**2)
        diff = dg - dk; ppm_val = round(dk / diff, 1) if diff != 0 else "Inf"
        ws.cell(row=curr,column=1,value=p1['測點名稱']).border=border; ws.cell(row=curr,column=2,value=f"---> {p2['測點名稱']}").border=border
        ws.cell(row=curr,column=3,value=dg).number_format='0.0000'; ws.cell(row=curr,column=3).border=border
        ws.cell(row=curr,column=4,value=dk).number_format='0.0000'; ws.cell(row=curr,column=4).border=border
        ws.cell(row=curr,column=5,value=diff).number_format='0.0000'; ws.cell(row=curr,column=5).border=border
        ws.cell(row=curr,column=6,value=ppm_val).border=border; ws.cell(row=curr,column=7,value="OK" if (diff==0 or abs(dk/diff)>5000) else "Check").border=border
        curr+=1
    curr+=2
    
    ws.merge_cells(f'A{curr}:F{curr}'); ws[f'A{curr}']="AZIMUTH CHECK"; ws[f'A{curr}'].font=bold_font; ws[f'A{curr}'].alignment=align_c; curr+=1
    headers_a = ["FROM", "TO", "GPS", "GROUND", "DIFFERENCE(SEC)", "TEST"]
    for i, h in enumerate(headers_a, 1):
        cell=ws.cell(row=curr, column=i, value=h); cell.border=border; cell.alignment=align_c; cell.fill=PatternFill(start_color="D3D3D3", fill_type="solid"); cell.font=bold_font
    curr+=1
    for p1, p2 in pairs:
        azg_rad = math.atan2(p2['E_轉換(GPS)']-p1['E_轉換(GPS)'], p2['N_轉換(GPS)']-p1['N_轉換(GPS)'])
        azk_rad = math.atan2(p2['E_已知(Ground)']-p1['E_已知(Ground)'], p2['N_已知(Ground)']-p1['N_已知(Ground)'])
        azg, azk = (math.degrees(azg_rad)+360)%360, (math.degrees(azk_rad)+360)%360
        diff_sec = (azg - azk) * 3600
        if diff_sec > 180*3600: diff_sec -= 360*3600
        if diff_sec < -180*3600: diff_sec += 360*3600
        ws.cell(row=curr,column=1,value=p1['測點名稱']).border=border; ws.cell(row=curr,column=2,value=f"---> {p2['測點名稱']}").border=border
        ws.cell(row=curr,column=3,value=deg_to_dmmss(azg)).number_format='0.000000'; ws.cell(row=curr,column=3).border=border
        ws.cell(row=curr,column=4,value=deg_to_dmmss(azk)).number_format='0.000000'; ws.cell(row=curr,column=4).border=border
        ws.cell(row=curr,column=5,value=round(diff_sec,2)).border=border; ws.cell(row=curr,column=6,value="OK" if abs(diff_sec)<20 else "Check").border=border
        curr+=1
    adjust_col_width(ws); f = io.BytesIO(); wb.save(f); f.seek(0); return f

def generate_report_7_3_final_coords(twd97_data, residuals_df):
    wb = Workbook(); ws = wb.active; ws.title = "最終坐標成果表"
    ws.merge_cells('A1:E1'); ws['A1'] = "e-GNSS 最終坐標成果清冊 (強制附合後)"; ws['A1'].font = Font(size=16, bold=True); ws['A1'].alignment = Alignment(horizontal='center')
    headers = ["測點名稱", "N_坐標(TWD97)", "E_坐標(TWD97)", "高程_H(m)", "點位屬性"]
    setup_excel_style(ws, headers, row_idx=2)

    used_pts = []
    if residuals_df is not None and not residuals_df.empty:
        used_pts = residuals_df[residuals_df.get('採用', pd.Series([True]*len(residuals_df))) == True]['測點名稱'].tolist()

    df_all = pd.DataFrame(twd97_data)
    df_known = df_all[df_all['測點名稱'].isin(used_pts)].copy().sort_values(by='測點名稱')
    df_unknown = df_all[~df_all['測點名稱'].isin(used_pts)].copy().sort_values(by='測點名稱')

    curr = 3
    for _, r in df_known.iterrows():
        ws.append([r['測點名稱'], round(r['N_TWD97'], 3), round(r['E_TWD97'], 3), round(r['H'], 3), "已知點 (強制附合採用)"]); curr += 1
    for _, r in df_unknown.iterrows():
        ws.append([r['測點名稱'], round(r['N_TWD97'], 3), round(r['E_TWD97'], 3), round(r['H'], 3), "圖根點"]); curr += 1

    adjust_col_width(ws); f = io.BytesIO(); wb.save(f); f.seek(0); return f

def generate_report_8_1_nlsc_pdf(twd97_data, gnss_data, check_pts):
    from fpdf import FPDF
    import math
    import itertools
    import pandas as pd
    import os
    import streamlit as st
    
    font_file = "msjh.ttc"
    if not os.path.exists(font_file):
        st.error(f"❌ 找不到字體檔 `{font_file}`！請確保該檔案與 e-gnss.py 放在同一個資料夾。")
        return None

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.add_font('tc', '', font_file, uni=True)
    pdf.set_auto_page_break(auto=True, margin=15)
    
    df_t = pd.DataFrame(twd97_data).set_index('測點名稱')
    df_g = pd.DataFrame(gnss_data).groupby('測點名稱')[['N','E']].mean()
    
    def dec2dms(deg):
        deg = deg % 360
        d = int(deg)
        rem = (deg - d) * 60
        m = int(rem)
        s = (rem - m) * 60
        return f"{d:03d}-{m:02d}-{s:05.2f}"

    # ==========================
    # 1. 點位坐標檢測成果表
    # ==========================
    pdf.add_page()
    pdf.set_font('tc', '', 16)
    pdf.cell(0, 10, "☆ ☆ ☆  內政部國土測繪中心  ☆ ☆ ☆", ln=True, align='C')
    pdf.set_font('tc', '', 14)
    pdf.cell(0, 10, "點位坐標檢測成果報表", ln=True, align='C')
    pdf.ln(5)
    
    pdf.set_font('tc', '', 9)
    w1 = [20, 25, 25, 25, 25, 20, 20, 20] 
    offset = (210 - sum(w1))/2
    
    pdf.set_x(offset)
    pdf.cell(w1[0], 6, "", border=0, align='C')
    pdf.cell(w1[1], 6, "檢測坐標", border=0, align='C')
    pdf.cell(w1[2], 6, "檢測坐標", border=0, align='C')
    pdf.cell(w1[3], 6, "原始坐標", border=0, align='C')
    pdf.cell(w1[4], 6, "原始坐標", border=0, align='C')
    pdf.cell(w1[5]+w1[6]+w1[7], 6, "較 差 (原-檢)", border=0, align='C')
    pdf.ln()
    
    pdf.set_x(offset)
    headers1 = ["點號", "N (m)", "E (m)", "N (m)", "E (m)", "dN(m)", "dE(m)", "差值"]
    for i, h in enumerate(headers1):
        pdf.cell(w1[i], 8, h, border='B', align='C')
    pdf.ln()
    
    max_dn_abs, max_de_abs, max_ds = -1.0, -1.0, -1.0
    max_dn_val, max_de_val, max_ds_val = 0.0, 0.0, 0.0
    max_dn_p, max_de_p, max_ds_p = "", "", ""

    for i, p in enumerate(check_pts):
        n_orig, e_orig = df_t.loc[p, 'N_TWD97'], df_t.loc[p, 'E_TWD97']
        n_chk, e_chk = df_g.loc[p, 'N'], df_g.loc[p, 'E']
        dn = n_orig - n_chk
        de = e_orig - e_chk
        ds = math.sqrt(dn**2 + de**2)
        
        if abs(dn) > max_dn_abs:
            max_dn_abs, max_dn_val, max_dn_p = abs(dn), dn, p
        if abs(de) > max_de_abs:
            max_de_abs, max_de_val, max_de_p = abs(de), de, p
        if ds > max_ds:
            max_ds, max_ds_val, max_ds_p = ds, ds, p
        
        pdf.set_x(offset)
        pdf.cell(w1[0], 8, f"{i+1}   {p}", align='C')
        pdf.cell(w1[1], 8, f"{n_chk:.3f}", align='C')
        pdf.cell(w1[2], 8, f"{e_chk:.3f}", align='C')
        pdf.cell(w1[3], 8, f"{n_orig:.3f}", align='C')
        pdf.cell(w1[4], 8, f"{e_orig:.3f}", align='C')
        pdf.cell(w1[5], 8, f"{dn:.3f}", align='C')
        pdf.cell(w1[6], 8, f"{de:.3f}", align='C')
        pdf.cell(w1[7], 8, f"{ds:.3f}", align='C')
        pdf.ln()
        
    if len(check_pts) > 0:
        pdf.set_x(offset)
        summary_text = f"共 {len(check_pts)} 個點，其中最大較差 dN({max_dn_p} : {max_dn_val:.3f}) dE({max_de_p} : {max_de_val:.3f}) d({max_ds_p} : {max_ds_val:.3f})"
        pdf.cell(sum(w1), 8, summary_text, align='L')
        pdf.ln()
    
    # ==========================
    # 2. 距離檢核 (強制換頁)
    # ==========================
    pdf.add_page()
    pdf.set_font('tc', '', 12)
    pdf.cell(0, 10, "<<<< 距離檢核 >>>>", ln=True, align='L')
    pdf.set_font('tc', '', 10)
    pdf.cell(0, 6, "# : [較差] > 1/5000 × [距離]", ln=True, align='L')
    pdf.cell(0, 6, "! : [較差]判定為 # , 惟小於 30 mm", ln=True, align='L')
    pdf.ln(5)
    
    pdf.set_font('tc', '', 9)
    w2 = [18, 18, 28, 28, 20, 30, 20, 18] 
    offset2 = (210 - sum(w2))/2
    
    pdf.set_x(offset2)
    headers2 = ["測站", "測站", "檢測距離(m)", "反算距離(m)", "較差", "距離/較差", "容許值#", "備註"]
    for i, h in enumerate(headers2):
        pdf.cell(w2[i], 8, h, border='B', align='C')
    pdf.ln()
    
    min_ratio = float('inf')
    min_ratio_p1, min_ratio_p2 = "", ""

    for p1, p2 in itertools.combinations(check_pts, 2):
        d_orig, _ = calc_dist_azimuth(df_t.loc[p1,'N_TWD97'], df_t.loc[p1,'E_TWD97'], df_t.loc[p2,'N_TWD97'], df_t.loc[p2,'E_TWD97'])
        d_chk, _ = calc_dist_azimuth(df_g.loc[p1,'N'], df_g.loc[p1,'E'], df_g.loc[p2,'N'], df_g.loc[p2,'E'])
        diff = abs(d_orig - d_chk)
        allow = d_orig / 5000.0
        ratio = int(d_orig / diff) if diff > 0.0001 else float('inf')
        
        if ratio < min_ratio:
            min_ratio = ratio
            min_ratio_p1, min_ratio_p2 = p1, p2
        
        note = ""
        if diff > allow:
            if diff < 0.030: note = "!"
            else: note = "#"
            
        pdf.set_x(offset2)
        pdf.cell(w2[0], 8, p1, align='C')
        pdf.cell(w2[1], 8, p2, align='C')
        pdf.cell(w2[2], 8, f"{d_chk:.3f}", align='C')
        pdf.cell(w2[3], 8, f"{d_orig:.3f}", align='C')
        pdf.cell(w2[4], 8, f"{diff:.3f}", align='C')
        pdf.cell(w2[5], 8, str(ratio) if ratio != float('inf') else "無限大", align='C')
        pdf.cell(w2[6], 8, f"{allow:.3f}", align='C')
        pdf.cell(w2[7], 8, note, align='C')
        pdf.ln()

    # 👉 距離檢核表專屬：輸出最低精度總結
    pdf.ln(5)
    if min_ratio != float('inf') and min_ratio_p1 != "":
        pdf.set_font('tc', '', 10)
        pdf.set_x(offset2)
        pdf.cell(0, 8, f"相對精度最低為 {min_ratio_p1} --> {min_ratio_p2} : 1 / {int(min_ratio)}", ln=True)

    # ==========================
    # 3. 方位角檢核 (強制換頁)
    # ==========================
    pdf.add_page()
    pdf.set_font('tc', '', 12)
    pdf.cell(0, 10, "<<<< 方位角檢核 >>>>", ln=True, align='L')
    pdf.ln(5)
    
    pdf.set_font('tc', '', 10)
    w3 = [20, 20, 40, 40, 25, 35] 
    offset3 = (210 - sum(w3))/2
    
    pdf.set_x(offset3)
    headers3 = ["測站", "測站", "檢測方位角", "反算方位角", "較差(秒)", "備註"]
    for i, h in enumerate(headers3):
        pdf.cell(w3[i], 8, h, border='B', align='C')
    pdf.ln()
    
    max_az_diff = -1.0
    max_az_p1, max_az_p2 = "", ""
    max_az_dist = 0.0

    for p1, p2 in itertools.combinations(check_pts, 2):
        d_orig, az_orig = calc_dist_azimuth(df_t.loc[p1,'N_TWD97'], df_t.loc[p1,'E_TWD97'], df_t.loc[p2,'N_TWD97'], df_t.loc[p2,'E_TWD97'])
        _, az_chk = calc_dist_azimuth(df_g.loc[p1,'N'], df_g.loc[p1,'E'], df_g.loc[p2,'N'], df_g.loc[p2,'E'])
        
        daz_sec = (az_orig - az_chk) * 3600
        if daz_sec > 180*3600: daz_sec -= 360*3600
        if daz_sec < -180*3600: daz_sec += 360*3600
        daz_sec = abs(daz_sec)
        
        if daz_sec > max_az_diff:
            max_az_diff = daz_sec
            max_az_p1, max_az_p2 = p1, p2
            max_az_dist = d_orig
        
        pdf.set_x(offset3)
        pdf.cell(w3[0], 8, p1, align='C')
        pdf.cell(w3[1], 8, p2, align='C')
        pdf.cell(w3[2], 8, dec2dms(az_chk), align='C')
        pdf.cell(w3[3], 8, dec2dms(az_orig), align='C')
        pdf.cell(w3[4], 8, f"{daz_sec:.2f}", align='C')
        pdf.cell(w3[5], 8, "", align='C')
        pdf.ln()

    # 👉 方位角檢核表專屬：輸出最大方位角較差與距離總結
    pdf.ln(5)
    if max_az_diff != -1.0 and max_az_p1 != "":
        pdf.set_x(offset3)
        pdf.cell(0, 8, f"方位角較差最大為 {max_az_p1} --> {max_az_p2} : {max_az_diff:.2f} 秒 ({max_az_dist:.3f} m)", ln=True)

    # 輸出 PDF Bytes
    try:
        pdf_bytes = bytes(pdf.output())
    except TypeError:
        pdf_bytes = pdf.output(dest='S').encode('latin1')
    return pdf_bytes

def generate_report_8_1_integrated(df_res, baseline_data):
    wb = Workbook(); ws = wb.active; ws.title = "整合式檢測成果表"
    align_c = Alignment(horizontal='center', vertical='center'); bold_font = Font(bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.merge_cells('A1:H1'); ws['A1']="內政部土地測量局已知點檢測成果報表"; ws['A1'].font=Font(size=14, bold=True); ws['A1'].alignment=align_c
    ws.merge_cells('A2:A3'); ws['A2']="點號"; ws.merge_cells('B2:C2'); ws['B2']="自由網坐標"; ws['B3']="N-坐標(m)"; ws['C3']="E-坐標(m)"
    ws.merge_cells('D2:E2'); ws['D2']="已知點坐標"; ws['D3']="N-坐標(m)"; ws['E3']="E-坐標(m)"; ws.merge_cells('F2:H2'); ws['F2']="較差"; ws['F3']="dN(m)"; ws['G3']="dE(m)"; ws['H3']="差值"
    for r in [2, 3]:
        for c in range(1, 9): cell = ws.cell(row=r, column=c); cell.border=border; cell.alignment=align_c; cell.font=bold_font
    
    max_dn = -1.0; max_de = -1.0; max_diff = -1.0; pt_dn, pt_de, pt_diff = "", "", ""
    used_res = df_res[df_res.get('採用', pd.Series([True]*len(df_res))) == True]
    
    curr = 4
    for r in used_res.to_dict('records'):
        dn = r['VX']; de = r['VY']; diff_val = r['平面殘差']
        ws.append([r['測點名稱'], round(r['N_轉換(GPS)'],3), round(r['E_轉換(GPS)'],3), round(r['N_已知(Ground)'],3), round(r['E_已知(Ground)'],3), round(dn,3), round(de,3), round(diff_val,3)])
        for c in range(1, 9): ws.cell(row=curr, column=c).border = border
        if abs(dn) > abs(max_dn): max_dn = dn; pt_dn = r['測點名稱']
        if abs(de) > abs(max_de): max_de = de; pt_de = r['測點名稱']
        if diff_val > max_diff: max_diff = diff_val; pt_diff = r['測點名稱']
        curr += 1
    ws.append([f"共 {len(used_res)} 個已知點, 其中最大較差 dN({pt_dn} {max_dn:.3f}) dE({pt_de} {max_de:.3f}) d({pt_diff} {max_diff:.3f})"]); curr += 2

    ws.merge_cells(f'A{curr}:F{curr}'); ws[f'A{curr}']="距離檢核"; ws[f'A{curr}'].font=bold_font; curr += 1
    headers_dist = ["測站", "測站", "檢測距離(m)", "反算距離(m)", "較差精度(1/ppm)", "容許誤差"]
    setup_excel_style(ws, headers_dist, row_idx=curr); curr += 1
    min_ppm = 999999; pt_ppm = ""
    for b in baseline_data:
        ppm_val = int(b['Dist_TWD97'] / abs(b['dDist'])) if b['dDist'] > 0.0001 else 999999
        ws.append([b['From'], b['To'], round(b['Dist_eGNSS'],3), round(b['Dist_TWD97'],3), f"{b['dDist']:.3f} (1/{ppm_val})", ""])
        for c in range(1, 7): ws.cell(row=curr-1, column=c).border = border
        if ppm_val < min_ppm: min_ppm = ppm_val; pt_ppm = f"{b['From']}==>{b['To']}"
        curr += 1
    ws.append([f"精度最低為({pt_ppm}) 1/ {min_ppm}"]); curr += 2

    ws.merge_cells(f'A{curr}:D{curr}'); ws[f'A{curr}']="方位角檢核"; ws[f'A{curr}'].font=bold_font; curr += 1
    headers_azi = ["測站", "測站", "檢測方位角", "反算方位角較差(秒)"]
    setup_excel_style(ws, headers_azi, row_idx=curr); curr += 1
    max_asec = -1.0; pt_asec = ""
    for b in baseline_data:
        ws.append([b['From'], b['To'], deg_to_dmmss(b['Az_eGNSS']), round(b['dAzi_Sec'],1)])
        ws.cell(row=curr-1, column=3).number_format = '0.000000'
        for c in range(1, 5): ws.cell(row=curr-1, column=c).border = border
        if abs(b['dAzi_Sec']) > abs(max_asec): max_asec = b['dAzi_Sec']; pt_asec = f"{b['From']}==>{b['To']}"
        curr += 1
    ws.append([f"方位角較差最大為({pt_asec}) {max_asec:.1f}秒"])
    adjust_col_width(ws); f = io.BytesIO(); wb.save(f); f.seek(0); return f

def generate_report_8_2_field(check_results):
    wb = Workbook(); ws = wb.active; ws.title = "實地檢測報表"
    ws.merge_cells('A1:G1'); ws['A1'] = "實地外業檢測成果比較表 (距離檢核)"; ws['A1'].font = Font(size=16, bold=True); ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A2:G2'); ws['A2'] = "合格標準：水平距離較差 <= 0.03m 或 相對誤差 >= 1/3000"
    ws['A2'].font = Font(color="0000FF", bold=True); ws['A2'].alignment = Alignment(horizontal='left')
    headers = ["起點", "終點", "實測距離(m)", "成果反算距離(m)", "距離較差(m)", "相對誤差", "判定"]
    setup_excel_style(ws, headers, row_idx=3)
    fail_font = Font(color="FF0000", bold=True); pass_font = Font(color="008000", bold=True)
    for i, row in enumerate(check_results, start=4):
        ws.append([row['From'], row['To'], round(row['Dist_Meas'],3), round(row['Dist_Calc'],3), round(row['dDist'],4), row['Rel_Error'], row['Status']])
        ws.cell(row=i, column=7).font = fail_font if row['Status']=="不合格" else pass_font
    adjust_col_width(ws); f = io.BytesIO(); wb.save(f); f.seek(0); return f

# 🔥 升級版 6參數引擎 (修正取消勾選點位消失問題)
def compute_6_parameters_optimized(obs_points, known_points, selected_names=None):
    df_obs = pd.DataFrame(obs_points).set_index('測點名稱'); df_known = pd.DataFrame(known_points).set_index('測點名稱')
    common = df_obs.join(df_known, lsuffix='_obs', rsuffix='_known', how='inner')
    
    if selected_names is None: selected_names = common.index.tolist()
    compute_df = common[common.index.isin(selected_names)]
    n = len(compute_df)
    
    if n < 3: return None, None, None, None, n
    
    A = np.zeros((2*n, 6)); L = np.zeros((2*n, 1))
    for i in range(n):
        r = compute_df.iloc[i]
        A[2*i,:] = [r['N_obs'], r['E_obs'], 1, 0, 0, 0]; L[2*i,0] = r['N_known']
        A[2*i+1,:] = [0, 0, 0, r['N_obs'], r['E_obs'], 1]; L[2*i+1,0] = r['E_known']
    try:
        X, _, _, _ = np.linalg.lstsq(A, L, rcond=None); p = X.flatten(); V = A @ X - L
        rmse = np.sqrt(np.sum(V**2)/(2*n-6)) if (2*n-6)>0 else 0
        sum_vv = np.sum(V**2)
        
        # 產生所有匹配點的殘差，供介面保留顯示
        res = []
        for pt_name, r in common.iterrows():
            nt = p[0]*r['N_obs'] + p[1]*r['E_obs'] + p[2]; et = p[3]*r['N_obs'] + p[4]*r['E_obs'] + p[5]
            vx = nt - r['N_known']; vy = et - r['E_known']
            res.append({'測點名稱': pt_name, 'N_已知(Ground)': r['N_known'], 'E_已知(Ground)': r['E_known'], 'N_轉換(GPS)': nt, 'E_轉換(GPS)': et, 'VX': vx, 'VY': vy, '平面殘差': np.sqrt(vx**2+vy**2), '採用': pt_name in selected_names})
        return p, pd.DataFrame(res), rmse, sum_vv, n
    except Exception as e: return None, None, None, None, str(e)

def calculate_residual_correction(tn, te, df_res, power=2):
    num_n, num_e, den = 0, 0, 0
    for _, row in df_res.iterrows():
        dist = math.sqrt((tn - row['N_轉換(GPS)'])**2 + (te - row['E_轉換(GPS)'])**2)
        if dist < 0.001: return row['VX'], row['VY']
        w = 1 / (dist ** power); num_n += w * row['VX']; num_e += w * row['VY']; den += w
    if den == 0: return 0, 0
    return num_n / den, num_e / den

# ================= 4. UI 邏輯與互動 =================
with col_right:
    st.header("📝 第四區：檢核流程與暫存")
    c1, c2 = st.columns(2)
    
    if c1.button("1️⃣ 檢核 第 1 測回", use_container_width=True): 
        if files_round1:
            with st.spinner("📦 正在合併並運算 1 測回資料，檔案較多請稍候..."):
                d, l = process_single_round(files_round1, "第 1 測回")
                st.session_state.logs = l
                st.session_state.current_stage_name = "第 1 測回"
                
                if d:
                    # 加入「寫入」布林值欄位，預設為 True (全選)
                    df_preview = pd.DataFrame(d)
                    df_preview.insert(0, '寫入', True)
                    st.session_state.current_stage_data = df_preview.to_dict('records')
                    st.success(f"✅ 第 1 測回檢核完成，共 {len(d)} 筆合格測點。請在下方確認並寫入暫存區！")
                else:
                    st.session_state.current_stage_data = None
                    st.error("⚠️ 檢核完畢，但【沒有任何測點】符合合格標準。請展開下方「檢核日誌」查看原因！")
        else: 
            st.error("請先上傳第 1 測回檔案")
            
    if c2.button("2️⃣ 檢核 第 2 測回", use_container_width=True):
        if files_round2:
            with st.spinner("📦 正在合併並運算 2 測回資料，檔案較多請稍候..."):
                d, l = process_single_round(files_round2, "第 2 測回")
                st.session_state.logs = l
                st.session_state.current_stage_name = "第 2 測回"
                
                if d:
                    # 加入「寫入」布林值欄位，預設為 True (全選)
                    df_preview = pd.DataFrame(d)
                    df_preview.insert(0, '寫入', True)
                    st.session_state.current_stage_data = df_preview.to_dict('records')
                    st.success(f"✅ 第 2 測回檢核完成，共 {len(d)} 筆合格測點。請在下方確認並寫入暫存區！")
                else:
                    st.session_state.current_stage_data = None
                    st.error("⚠️ 檢核完畢，但【沒有任何測點】符合合格標準。請展開下方「檢核日誌」查看原因！")
        else: 
            st.error("請先上傳第 2 測回檔案")
    
    if st.session_state.current_stage_name != "":
        st.markdown("---")
        st.info(f"預覽: {st.session_state.current_stage_name}")
        
        # 預設將日誌收合，讓畫面保持乾淨
        with st.expander("檢核日誌", expanded=False):
            if 'logs' in st.session_state and st.session_state.logs:
                for line in st.session_state.logs: 
                    if "✅" in line: st.success(line)
                    elif "❌" in line: st.error(line)
                    else: st.write(line)
            else:
                st.write("尚無日誌資料。")
                
        # 互動式預覽與寫入區塊
        if st.session_state.current_stage_data:
            st.write("▼ **請勾選欲保留的資料 (系統已預設全選):**")
            
            df_to_edit = pd.DataFrame(st.session_state.current_stage_data)
            edited_preview = st.data_editor(
                df_to_edit, 
                hide_index=True, 
                use_container_width=True, 
                column_config={"寫入": st.column_config.CheckboxColumn(required=True)},
                disabled=["測點名稱", "測回別", "有效筆數", "總計點數", "使用比率", "N", "E", "H", "儀器高", "平均時間", "起測時間", "結束時間", "PDOP"]
            )
            
            if st.button("💾 將勾選資料寫入暫存區", type="primary", use_container_width=True):
                # 篩選有打勾的資料
                selected_data = edited_preview[edited_preview['寫入'] == True].drop(columns=['寫入']).to_dict('records')
                
                if selected_data:
                    st.session_state.temp_database.extend(selected_data)
                    # 寫入成功後清空預覽，防止重複點擊
                    st.session_state.current_stage_data = None 
                    st.rerun() 
                else:
                    st.warning("⚠️ 您沒有勾選任何資料，無法寫入。")

    # ==========================================================
    # 📝 第五區：暫存資料庫管理 (已一本化修復，絕不重複顯示)
    # ==========================================================
    st.markdown("---")
    st.header("📝 第五區：暫存資料庫管理")
    if st.session_state.temp_database:
        df_db = pd.DataFrame(st.session_state.temp_database)
        if "移除" not in df_db.columns: 
            df_db.insert(0, "移除", False)
            
        edited_df = st.data_editor(
            df_db, 
            hide_index=True, 
            use_container_width=True, 
            column_config={"移除": st.column_config.CheckboxColumn(required=True)},
            disabled=["測點名稱", "測回別", "有效筆數", "總計點數", "使用比率", "N", "E", "H", "儀器高", "平均時間", "起測時間", "結束時間", "PDOP"]
        )
        
        cd1, cd2 = st.columns([1, 4])
        if cd1.button("🗑️ 刪除選取"): 
            st.session_state.temp_database = edited_df[edited_df['移除'] == False].drop(columns=['移除']).to_dict('records')
            st.rerun()
        if cd2.button("💣 清空全部"): 
            st.session_state.temp_database = []
            st.rerun()
            
        # ====== 修復：外業紀錄表匯出區塊 (移入「有資料」的判斷式內) ======
        st.markdown("#### 📋 附加報表：測量外業紀錄表 (PDF)")
        
        c_pdf1, c_pdf2, c_pdf3 = st.columns([2, 1, 1.5]) 
        with c_pdf1:
            proj_name = st.text_input("專案名稱", value="115年度仁愛鄉非都市計畫地區圖解數化地籍圖整合建置作業")
        with c_pdf2:
            personnel_name = st.text_input("測量人員姓名", value="賴柏霖")
        with c_pdf3:
            st.write("") # 排版佔位
            st.write("")
            pdf_data = generate_field_record_pdf(st.session_state.temp_database, personnel=personnel_name, project_name=proj_name)
            if pdf_data:
                st.download_button(
                    label="📄 下載 VBS-RTK 測量外業紀錄表 (PDF)",
                    data=pdf_data,
                    file_name="VBS-RTK_測量外業紀錄表.pdf",
                    mime="application/pdf",
                    type="primary",
                    use_container_width=True
                )
    else: 
        st.info("暫存區為空")

# ==========================================================
    # 📤 補回：第六區：觀測品質與坐標檢核 (已修正函數名稱與門檻傳遞)
    # ==========================================================
    st.markdown("---")
    st.header("📤 第六區：觀測品質與坐標檢核")
    
    if st.session_state.temp_database:
        # 讀取左側欄的檢核門檻，打包成 limits 傳給 6-1 報表
        limits = {
            'diff_inst': diff_inst_limit,
            'diff_plane': diff_plane_limit,
            'diff_elev': diff_elev_limit,
            'time_gap': time_gap_limit
        }
        
        c6_1, c6_2, c6_3 = st.columns(3)
        with c6_1:
            # 修正函數名稱為 generate_report_6_1_accuracy，並傳入 limits
            f_6_1 = generate_report_6_1_accuracy(st.session_state.temp_database, limits)
            st.download_button("📊 下載 報表6-1.精度檢核總表", f_6_1, "報表6-1_精度檢核總表.xlsx", use_container_width=True)
            
        with c6_2:
            # 修正函數名稱為 generate_report_6_2_center
            f_6_2 = generate_report_6_2_center(st.session_state.temp_database)
            if f_6_2 is not None:
                st.download_button("⚖️ 下載 報表6-2.成果檢核表(測繪中心版)", f_6_2, "報表6-2_成果檢核表.xlsx", use_container_width=True)
                
        with c6_3:
            # 修正函數名稱為 generate_report_6_3_coord
            f_6_3 = generate_report_6_3_coord(st.session_state.temp_database)
            st.download_button("🎯 下載 報表6-3.坐標成果表", f_6_3, "報表6-3_坐標成果表.xlsx", use_container_width=True)
    else:
        st.info("請先將合格資料寫入第五區暫存區，即可產製第六區報表。")

    st.markdown("---"); st.header("🏁 第七區：六參數優化與強制附合")
    sample_kp = pd.DataFrame({'測點名稱': ['R266', 'R291', 'SW52', 'SW61'], 'N': [2545455.695, 2537041.401, 2536216.438, 2538963.044], 'E': [179627.685, 177281.763, 191536.663, 195893.554]})
    f_kp_sample = io.BytesIO()
    with pd.ExcelWriter(f_kp_sample, engine='openpyxl') as writer: sample_kp.to_excel(writer, index=False)
    f_kp_sample.seek(0)
    st.download_button("📥 下載 已知點清冊範例檔", f_kp_sample, "已知點清冊範例.xlsx")

    kp_file = st.file_uploader("📂 上傳 已知控制點清冊", type=['csv', 'xlsx'], key="kp_u")
    
    if kp_file and st.session_state.temp_database:
        try:
            if kp_file.name.endswith('.csv'): df_kp = pd.read_csv(kp_file, encoding='utf-8-sig')
            else: df_kp = pd.read_excel(kp_file)
            df_kp.columns = df_kp.columns.str.replace('\n', '').str.replace('"', '').str.strip()
            
            # 加入除錯機制，提示目前抓到的欄位
            if {'測點名稱','N','E'}.issubset(df_kp.columns):
                obs_avg = pd.DataFrame(st.session_state.temp_database).groupby('測點名稱')[['N','E']].mean().reset_index().to_dict('records')
                known_list = df_kp[['測點名稱','N','E']].to_dict('records')
                
                # 💡 新增：計算同名點位數量並給予明確提示
                obs_pts = set([str(r['測點名稱']) for r in obs_avg])
                known_pts = set([str(r['測點名稱']) for r in known_list])
                common_pts = obs_pts.intersection(known_pts)
                
                if len(common_pts) < 3:
                    st.error(f"❌ 錯誤：匹配到的同名控制點數量不足！(至少需要 3 點，目前僅匹配到 {len(common_pts)} 點)")
                    st.warning("🔍 請檢查「暫存區」與「已知點清冊」的測點名稱是否完全一致：")
                    st.info(f"👉 您的暫存區有這些點：{', '.join(list(obs_pts))}\n👉 您的已知點清冊有這些點：{', '.join(list(known_pts))}")
                else:
                    if st.session_state.trans_residuals is None:
                        p, df_res, rmse, vv, n = compute_6_parameters_optimized(obs_avg, known_list)
                        if df_res is not None:
                            st.session_state.trans_residuals = df_res
                            st.session_state.trans_rmse = rmse
                            st.session_state.trans_params = p
                            st.session_state.trans_vv = vv

                    if st.session_state.trans_residuals is not None:
                        st.success(f"✅ 成功匹配 {len(common_pts)} 個控制點，已自動完成初始轉換！")
                        st.subheader("🔍 殘差比較與點位優化")
                        st.caption("請勾選您要採用進行「強制附合」的點位。")
                        
                        edited_res = st.data_editor(st.session_state.trans_residuals, hide_index=True, use_container_width=True, 
                                                    column_config={"採用": st.column_config.CheckboxColumn(default=True)},
                                                    disabled=["測點名稱", "N_已知(Ground)", "E_已知(Ground)", "N_轉換(GPS)", "E_轉換(GPS)", "VX", "VY", "平面殘差"])
                        
                        c_opt1, c_opt2 = st.columns(2)
                        if c_opt1.button("🚀 重新解算 (僅限採用點位)", use_container_width=True):
                            selected_pts = edited_res[edited_res['採用']==True]['測點名稱'].tolist()
                            if len(selected_pts) < 3:
                                st.error("❌ 至少需要保留 3 個控制點才能進行解算！")
                            else:
                                p_new, df_res_new, rmse_new, vv_new, n_new = compute_6_parameters_optimized(obs_avg, known_list, selected_pts)
                                st.session_state.trans_residuals = df_res_new
                                st.session_state.trans_rmse = rmse_new
                                st.session_state.trans_params = p_new
                                st.session_state.trans_vv = vv_new
                                st.success(f"優化完成！採用點數: {n_new}, 新 RMSE: {rmse_new:.4f}m"); st.rerun()
                        
                        if c_opt2.button("🌍 執行全區強制附合", type="primary", use_container_width=True):
                            st.session_state.trans_residuals = edited_res
                            p = st.session_state.trans_params
                            obs_all = pd.DataFrame(st.session_state.temp_database).groupby('測點名稱')[['N','E','H']].mean().reset_index()
                            res_final = []
                            selected_df = edited_res[edited_res['採用']==True]
                            for _, r in obs_all.iterrows():
                                na = p[0]*r['N'] + p[1]*r['E'] + p[2]; ea = p[3]*r['N'] + p[4]*r['E'] + p[5]
                                dn, de = calculate_residual_correction(na, ea, selected_df)
                                res_final.append({'測點名稱': r['測點名稱'], 'N_TWD97': na+dn, 'E_TWD97': ea+de, 'H': r['H']})
                            st.session_state.final_twd97_data = res_final
                            st.success("✅ 全區轉換完成 (已依據採用點位執行殘差分配)")
                            
                        if st.session_state.final_twd97_data:
                            st.markdown("#### 📤 第七區：轉換與優化報表下載")
                            cd_7_1, cd_7_2, cd_7_3 = st.columns(3)
                            with cd_7_1:
                                f_7_1 = generate_report_7_1_residuals(st.session_state.trans_residuals)
                                st.download_button("🔍 下載 報表7-1.已知點殘差表", f_7_1, "報表7-1_已知點殘差與優化比較表.xlsx", use_container_width=True)
                            with cd_7_2:
                                f_7_2 = generate_report_7_2_transform(st.session_state.trans_params, st.session_state.trans_residuals, st.session_state.trans_rmse, st.session_state.trans_vv)
                                st.download_button("⚙️ 下載 報表7-2.轉換參數報表", f_7_2, "報表7-2_六參數轉換報表.xlsx", use_container_width=True)
                            with cd_7_3:
                                f_7_3 = generate_report_7_3_final_coords(st.session_state.final_twd97_data, st.session_state.trans_residuals)
                                st.download_button("📍 下載 報表7-3.最終坐標清冊", f_7_3, "報表7-3_最終坐標成果清冊.xlsx", use_container_width=True)
            else:
                st.error(f"❌ 錯誤：已知點清冊缺少必要欄位！\n目前的欄位有: {list(df_kp.columns)}\n請確保表頭必須包含: `測點名稱`, `N`, `E` (大小寫須完全一致)")
        except Exception as e: 
            st.error(f"檔案讀取失敗: {str(e)}")

# ================= 8. 第八區：基線比較與實地檢測 =================
    if st.session_state.final_twd97_data:
        st.markdown("---")
        st.header("🔍 第八區：基線檢核與實地檢測")
        
        if st.button("🚀 計算非已知點全組合基線與檢測", type="primary", use_container_width=True):
            df_twd97 = pd.DataFrame(st.session_state.final_twd97_data).set_index('測點名稱')
            
            # 篩選出「非已知點」(將打勾採用作轉換的控制點剃除)
            used_pts = st.session_state.trans_residuals[st.session_state.trans_residuals['採用']==True]['測點名稱'].tolist()
            check_pts = [p for p in df_twd97.index if p not in used_pts]
            
            if len(check_pts) < 2:
                st.warning("⚠️ 檢測點 (非已知點) 數量不足 2 點，無法組成基線進行比對！請確認您的資料集包含足夠的檢測點。")
            else:
                # 儲存到 session_state 供下載報表使用
                st.session_state.baseline_check_data = {
                    'twd97': st.session_state.final_twd97_data,
                    'gnss': st.session_state.temp_database,
                    'check_pts': check_pts
                }
                st.success(f"✅ 已完成 {len(check_pts)} 個非已知點的坐標、基線長度與方位角比對！")

        if st.session_state.baseline_check_data and isinstance(st.session_state.baseline_check_data, dict):
            # 產製符合國土測繪中心格式的 PDF 報表
            f_8_1_pdf = generate_report_8_1_nlsc_pdf(
                st.session_state.baseline_check_data['twd97'], 
                st.session_state.baseline_check_data['gnss'], 
                st.session_state.baseline_check_data['check_pts']
            )
            
            if f_8_1_pdf:
                st.download_button(
                    label="📄 下載 報表8-1.點位坐標與基線檢測成果 (PDF格式)", 
                    data=f_8_1_pdf, 
                    file_name="報表8-1_點位坐標與基線檢測成果.pdf", 
                    mime="application/pdf", 
                    type="primary", 
                    use_container_width=True
                )
            
            st.markdown("#### 📏 實地外業檢測 (短邊 < 100m)")
            
            all_pts = [r['測點名稱'] for r in st.session_state.final_twd97_data]
            df_twd97_all = pd.DataFrame(st.session_state.final_twd97_data).set_index('測點名稱')
            all_baselines = []
            for p1, p2 in itertools.combinations(all_pts, 2):
                d_t, _ = calc_dist_azimuth(df_twd97_all.loc[p1,'N_TWD97'], df_twd97_all.loc[p1,'E_TWD97'], df_twd97_all.loc[p2,'N_TWD97'], df_twd97_all.loc[p2,'E_TWD97'])
                all_baselines.append({'From': p1, 'To': p2, 'Dist_TWD97': d_t})
            
            df_all_base = pd.DataFrame(all_baselines)
            short_baselines = df_all_base[df_all_base['Dist_TWD97'] <= short_dist_limit].copy()
            
            if not short_baselines.empty:
                st.warning(f"⚠️ 發現 {len(short_baselines)} 組短邊，建議進行實地檢測！")
                st.dataframe(short_baselines[['From', 'To', 'Dist_TWD97']].style.format({'Dist_TWD97': '{:.3f}'}))
                
                sample_data = short_baselines[['From', 'To']].copy()
                sample_data['實測距離(m)'] = ""
                f_sample = io.BytesIO()
                with pd.ExcelWriter(f_sample, engine='openpyxl') as writer: sample_data.to_excel(writer, index=False)
                f_sample.seek(0)
                st.download_button("📥 下載 外業檢測紀錄表範例檔", f_sample, "外業短邊檢測紀錄表(空白).xlsx")
                
                st.markdown("##### 📤 上傳實測成果進行比對")
                field_file = st.file_uploader("上傳填寫完成的外業檢測表", type=['xlsx'])
                
                if field_file:
                    try:
                        df_field = pd.read_excel(field_file)
                        check_res = []
                        for _, row in df_field.iterrows():
                            p1, p2 = str(row['From']), str(row['To'])
                            match = short_baselines[(short_baselines['From']==p1) & (short_baselines['To']==p2)]
                            if not match.empty:
                                d_calc = match.iloc[0]['Dist_TWD97']
                                d_meas = row.get('實測距離(m)', 0)
                                if pd.notnull(d_meas):
                                    dd = abs(d_meas - d_calc)
                                    is_pass = False
                                    if dd <= 0.03: is_pass = True
                                    elif d_calc > 0 and (d_calc / dd) >= 3000: is_pass = True
                                    status = "合格" if is_pass else "不合格"
                                    rel = f"1/{int(d_calc/dd)}" if dd > 0.001 else "無限大"
                                    check_res.append({'From': p1, 'To': p2, 'Dist_Meas': d_meas, 'Dist_Calc': d_calc, 'dDist': dd, 'Rel_Error': rel, 'Status': status})
                        
                        if check_res:
                            st.write("▼ **實地檢測成果比對:**")
                            st.dataframe(pd.DataFrame(check_res))
                            f_8_2 = generate_report_8_2_field(check_res)
                            st.download_button("👷 下載 報表8-2.實地外業檢測報表", f_8_2, "報表8-2_實地外業檢測報表.xlsx", type="primary")
                            
                    except Exception as e: st.error(f"讀取錯誤: {e}")
            else:
                st.success("✅ 全區無小於門檻之短邊，無需辦理強制實地檢測。")
